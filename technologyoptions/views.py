import math
from io import BytesIO

import numpy as np
import openpyxl
import pandas as pd
import xlwt
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import redirect
from django.shortcuts import render
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.utils.translation import gettext_lazy as _
from django.views.generic import CreateView
from django.views.generic import TemplateView
from django.views.generic.edit import DeleteView
from django.views.generic.edit import UpdateView
from django.views.generic.list import ListView
from quotations.models import QuotationIntroPara

from technologyoptions.forms import TechnologyOptionCreateForm
from technologyoptions.forms import TechnologySubComponentCreateForm
from technologyoptions.forms import TechnologySubComponentUpdateForm
from technologyoptions.forms import TechnologyTypeCreateForm
from technologyoptions.forms import TechnologyTypeUpdateForm
from technologyoptions.models import TechnologyOption
from technologyoptions.models import TechnologySubComponent
from technologyoptions.models import TechnologyType


@method_decorator(login_required, name="dispatch")
class OnlyForAdmin:
    def dispatch(self, request, *args, **kwargs):

        if not request.user.is_superuser:
            return redirect(reverse_lazy("quotations:quotations"))

        return super().dispatch(request, *args, **kwargs)


class AddTechnoTypeView(OnlyForAdmin, CreateView):
    template_name = "add_techno_type.html"
    form_class = TechnologyTypeCreateForm
    success_url = reverse_lazy("technologyoptions:add-techno-type")

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context["technology_types"] = TechnologyType.objects.all()
        return context


class UpdateTechnoTypeView(OnlyForAdmin, UpdateView):
    model = TechnologyType
    template_name = "update_techno_type.html"
    form_class = TechnologyTypeUpdateForm
    success_url = reverse_lazy("technologyoptions:add-techno-type")


class TechnologyTypeDeleteView(OnlyForAdmin, DeleteView):
    model = TechnologyType
    template_name = "delete_techno_type.html"

    def get(self, request, pk, *args, **kwargs):
        self.object = self.get_object()
        context = self.get_context_data(object=self.object)
        technology_type = TechnologyType.objects.filter(pk=pk).first()
        technology_option = TechnologyOption.objects.filter(
            technology_type=technology_type
        ).first()
        technology_sub_component = TechnologySubComponent.objects.filter(
            technology_type=technology_type
        ).first()
        intropara = QuotationIntroPara.objects.filter(
            technology_type=technology_type
        ).first()
        if technology_option or technology_sub_component or intropara:
            messages.error(
                request,
                f"Cannot delete technology / service",
            )
            if technology_option:
                context["technology_options"] = TechnologyOption.objects.filter(
                    technology_type=technology_type
                )
            if technology_sub_component:
                context[
                    "technology_sub_components"
                ] = TechnologySubComponent.objects.filter(
                    technology_type=technology_type
                )
            if intropara:
                context["intro_paras"] = QuotationIntroPara.objects.filter(
                    technology_type=technology_type
                )
            context["pk"] = pk
        return render(request, "delete_techno_type.html", context)

    def post(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.delete()
        return redirect(reverse_lazy("technologyoptions:add-techno-type"))


class TechnologyOptionListView(OnlyForAdmin, ListView):
    model = TechnologyOption
    template_name = "technology_options.html"
    success_url = reverse_lazy("technologyoptions:product-or-part-numbers")

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset


class AddTechnologyOptionView(OnlyForAdmin, CreateView):
    model = TechnologyOption
    template_name = "add_technology_option.html"
    form_class = TechnologyOptionCreateForm
    success_url = reverse_lazy("technologyoptions:product-or-part-numbers")


class UpdateTechnologyOptionView(OnlyForAdmin, UpdateView):
    model = TechnologyOption
    template_name = "update_technology_option.html"
    form_class = TechnologyOptionCreateForm
    success_url = reverse_lazy("technologyoptions:product-or-part-numbers")


class TechnologyOptionDeleteView(OnlyForAdmin, DeleteView):
    model = TechnologyOption
    template_name = "delete_technology_option.html"
    success_url = reverse_lazy("technologyoptions:product-or-part-numbers")


class BulkUploadTechnologyOptionsView(OnlyForAdmin, TemplateView):
    template_name = "bulk_upload_technology_options.html"
    success_url = reverse_lazy("technologyoptions:product-or-part-numbers")

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES["bulk_technology_file"]
        category = request.POST.get("category")
        read_file = pd.read_excel(excel_file)

        """
        If Sheet is empty , throw error message ."""
        if read_file.empty:
            messages.error(
                request,
                f"Failed bulk upload: Selected file is empty. Please fill-up the valid data in the Excel File.   ",
            )
            return redirect(
                reverse_lazy("technologyoptions:bulk-upload-products-or-part-numbers")
            )

        wb = openpyxl.load_workbook(excel_file)
        worksheets = wb.sheetnames
        ps_cost_validation_index_dict, fields_dict = {}, {}
        valid_columns_list = [
            "TECHNOLOGY / SERVICE",
            "TECHNOLOGY / SERVICE COMPONENT",
            "TECHNOLOGY / SERVICE SUB-COMPONENT",
            "PART NUMBER",
            "DESCRIPTION",
            "UNIT",
            "PS COST",
            "DISTRIBUTION COST",
            "END USER COST",
            "CATEGORY",
        ]

        """
        when file template has less columns than the valid columns_list, 
        then data will be saved in mismatched order in database, to avoid that, 
        this validation error is needed."""
        for worksheet in worksheets:
            excel_data = pd.read_excel(excel_file, worksheet)
            columns_names = excel_data.columns.values.tolist()

            if "QUESTIONS" in columns_names:
                columns_names.pop(columns_names.index("QUESTIONS"))

            fields_list = [
                column for column in valid_columns_list if column not in columns_names
            ]
            if fields_list:
                fields_dict[worksheet] = fields_list

        if fields_dict:
            nl = "\n"
            messages.error(
                request,
                f"Failed bulk upload: mentioned columns should be listed out in respective worksheets :{nl}{nl.join(f'{key}: {value}' for key,value in fields_dict.items())}",
            )
            return redirect(
                reverse_lazy("technologyoptions:bulk-upload-products-or-part-numbers")
            )

        """
        when in sheet if ps_cost is greater than distribution cost or enduser cost, 
        then this error message will be throw."""

        ps_cost_validation_index_dict, null_data_index_dict = {}, {}
        for worksheet in worksheets:
            excel_data = pd.read_excel(excel_file, worksheet)
            index_list, index_list1 = [], []
            for index in excel_data.index:
                ps_cost = excel_data["PS COST"][index]
                distribution_cost = excel_data["DISTRIBUTION COST"][index]
                end_user_cost = excel_data["END USER COST"][index]

                if (
                    isinstance(excel_data["PS COST"][index], str)
                    or isinstance(excel_data["DISTRIBUTION COST"][index], str)
                ) or isinstance(excel_data["END USER COST"][index], str):
                    continue

                excel_data["PS COST"][index] = ps_cost
                excel_data["DISTRIBUTION COST"][index] = distribution_cost
                excel_data["END USER COST"][index] = end_user_cost

                if (math.isnan(ps_cost) or math.isnan(distribution_cost)) or math.isnan(
                    end_user_cost
                ):
                    index_list1.append(index + 2)
                if (
                    excel_data["PS COST"][index]
                    >= excel_data["DISTRIBUTION COST"][index]
                    or excel_data["PS COST"][index]
                    >= excel_data["END USER COST"][index]
                ):
                    index_list.append(index + 2)

            if index_list1:
                null_data_index_dict[worksheet] = index_list1
            if index_list:
                ps_cost_validation_index_dict[worksheet] = index_list

        if ps_cost_validation_index_dict or null_data_index_dict:
            nl = "\n"
            if ps_cost_validation_index_dict:
                messages.error(
                    request,
                    f"""Failed bulk upload: PS cost should be less than distribution cost and end user cost at index in respective worksheets:{nl}
                    {nl.join(f'{key}: {value}' for key,value in ps_cost_validation_index_dict.items())}
                    """,
                )
            if null_data_index_dict:
                messages.error(
                    request,
                    f"""Failed bulk upload: Data is missing at index in respective worksheets:{nl}
                {nl.join(f'{key}: {value}' for key,value in null_data_index_dict.items())}
                """,
                )
            return redirect(
                reverse_lazy("technologyoptions:bulk-upload-products-or-part-numbers")
            )
        else:
            # iterating over the rows and
            # getting value from each cell in row
            for worksheet in worksheets:
                for index, row in enumerate(wb[worksheet].iter_rows()):
                    if index == 0:
                        continue

                    row_data = list()
                    for cell in row:
                        row_data.append(cell.value)

                    if not row_data[0] and not row_data[1]:
                        break

                    technology_type = TechnologyType.objects.filter(
                        type__iexact=row_data[0]
                    ).first()

                    if not technology_type:
                        technology_type = TechnologyType.objects.create(
                            type=row_data[0]
                        )
                        technology_type.save()

                    technology_component = row_data[1].title()

                    sub_component = TechnologySubComponent.objects.filter(
                        sublist_item_name__iexact=row_data[2],
                        technology_component__iexact=technology_component,
                        technology_type=technology_type,
                    ).first()

                    if not sub_component:
                        sub_component = TechnologySubComponent.objects.create(
                            technology_type=TechnologyType.objects.get(
                                type=row_data[0]
                            ),
                            technology_component=technology_component,
                            sublist_item_name=row_data[2],
                        )
                        sub_component.save()

                    part_number = row_data[3]
                    description = row_data[4]
                    unit = row_data[5].title()
                    ps_cost = (
                        round(0, 2) if isinstance(row_data[6], str) else row_data[6]
                    )
                    distribution_cost = (
                        round(0, 2) if isinstance(row_data[7], str) else row_data[7]
                    )
                    end_user_cost = (
                        round(0, 2) if isinstance(row_data[8], str) else row_data[8]
                    )
                    category = row_data[9]

                    technology_option_queryset = TechnologyOption.objects.filter(
                        technology_type=technology_type,
                        technology_component=technology_component,
                        technology_sub_component=sub_component,
                        part_number__iexact=row_data[3],
                        category = category,
                    )
                    if technology_option_queryset.first():
                        technology_option_queryset.update(
                            technology_type=technology_type,
                            technology_component=technology_component,
                            technology_sub_component=sub_component,
                            part_number=part_number,
                            description=description,
                            unit=unit,
                            ps_cost=ps_cost,
                            distribution_cost=distribution_cost,
                            end_user_cost=end_user_cost,
                            category=category,
                        )
                    else:
                        TechnologyOption.objects.create(
                            technology_type=technology_type,
                            technology_component=technology_component,
                            technology_sub_component=sub_component,
                            part_number=part_number,
                            description=description,
                            unit=unit,
                            ps_cost=ps_cost,
                            distribution_cost=distribution_cost,
                            end_user_cost=end_user_cost,
                            category=category,
                        )

        return redirect(reverse_lazy("technologyoptions:product-or-part-numbers"))


def technology_sub_component_list(request, **kwargs):
    technology_sub_component_queryset = TechnologySubComponent.objects.filter(
        technology_component=kwargs["technology_component"],
        technology_type=TechnologyType.objects.get(id=kwargs["technology_type"]),
    )

    data = {
        sub_component.id: sub_component.sublist_item_name
        for sub_component in technology_sub_component_queryset
    }

    return JsonResponse(data)


def bulk_upload_template(request):
    excelfile = BytesIO()
    filename = "Bulk Products or Part Numbers"
    book = xlwt.Workbook(excelfile)
    sheet = book.add_sheet("Sheet")
    row = 0
    column = 0
    columns = [
        "TECHNOLOGY / SERVICE",
        "TECHNOLOGY / SERVICE COMPONENT",
        "TECHNOLOGY / SERVICE SUB-COMPONENT",
        "PART NUMBER",
        "DESCRIPTION",
        "UNIT",
        "PS COST",
        "DISTRIBUTION COST",
        "END USER COST",
    ]
    style_string = "font: bold on;"
    style = xlwt.easyxf(style_string)
    for col in columns:
        sheet.write(row, column, col, style=style)
        column += 1
    book.save(excelfile)
    response = HttpResponse(
        excelfile.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename={filename}.xlsx"
    return response


class TechnologySubcomponentCreateView(OnlyForAdmin, CreateView):
    model = TechnologySubComponent
    template_name = "add_technology_subcomponent.html"
    form_class = TechnologySubComponentCreateForm
    success_url = reverse_lazy("technologyoptions:add-technology-subcomponent")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["technololgy_subcomponents"] = TechnologySubComponent.objects.all()
        return context


class TechnologySubcomponentUpdateView(OnlyForAdmin, UpdateView):
    model = TechnologySubComponent
    template_name = "update_technology_subcomponent.html"
    form_class = TechnologySubComponentUpdateForm
    success_url = reverse_lazy("technologyoptions:add-technology-subcomponent")


class TechnologySubcomponentDeleteView(OnlyForAdmin, DeleteView):
    model = TechnologySubComponent
    template_name = "delete_technology_subcomponent.html"
    success_url = reverse_lazy("technologyoptions:add-technology-subcomponent")

    def get(self, request, pk, *args, **kwargs):
        self.object = self.get_object()
        context = self.get_context_data(object=self.object)
        technology_sub_component = TechnologySubComponent.objects.filter(pk=pk).first()
        technology_option = TechnologyOption.objects.filter(
            technology_sub_component=technology_sub_component
        ).first()
        if technology_option:
            messages.error(
                request,
                f"Cannot delete technology / service",
            )
            context["technology_options"] = TechnologyOption.objects.filter(
                technology_sub_component=technology_sub_component
            )
            context["pk"] = pk
        return render(request, "delete_technology_subcomponent.html", context)
